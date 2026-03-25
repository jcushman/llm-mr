from __future__ import annotations

import csv
import json
from contextlib import contextmanager
from pathlib import Path
from typing import IO, Iterable, Iterator, Sequence

from openpyxl import Workbook, load_workbook

from .registries import InputRegistry, OutputRegistry, Row, RowAppender, TableStream


class CSVInputPlugin:
    name = "csv"
    extensions = [".csv"]
    typed = False

    @contextmanager
    def open(self, path: Path) -> Iterator[TableStream]:
        with path.open("r", encoding="utf-8", newline="") as fp:
            yield self._read(fp)

    @contextmanager
    def open_stream(self, stream: IO[str]) -> Iterator[TableStream]:
        yield self._read(stream)

    @staticmethod
    def _read(fp: IO[str]) -> TableStream:
        reader = csv.DictReader(fp)
        fieldnames = list(reader.fieldnames or [])
        return TableStream(rows=reader, fieldnames=fieldnames)


class _CSVRowAppender:
    def __init__(self, fp: IO[str], fieldnames: list):
        self._fp = fp
        self._writer = csv.DictWriter(fp, fieldnames=fieldnames)

    def append(self, row: Row) -> None:
        self._writer.writerow(row)

    def flush(self) -> None:
        self._fp.flush()


class CSVOutputPlugin:
    name = "csv"
    extensions = [".csv"]

    def write(self, path: Path, rows: Iterable[Row], fieldnames: Iterable[str]) -> None:
        field_list = list(fieldnames)
        with path.open("w", encoding="utf-8", newline="") as fp:
            self._write_fp(fp, rows, field_list)

    def write_stream(
        self, stream: IO[str], rows: Iterable[Row], fieldnames: Iterable[str]
    ) -> None:
        self._write_fp(stream, rows, list(fieldnames))

    @contextmanager
    def open_append(
        self, path: Path, fieldnames: Sequence[str]
    ) -> Iterator[RowAppender]:
        field_list = list(fieldnames)
        write_header = not path.exists() or path.stat().st_size == 0
        fp = path.open("a", encoding="utf-8", newline="")
        try:
            if write_header and field_list:
                writer = csv.DictWriter(fp, fieldnames=field_list)
                writer.writeheader()
                fp.flush()
            yield _CSVRowAppender(fp, field_list)
        finally:
            fp.close()

    @staticmethod
    def _write_fp(fp: IO[str], rows: Iterable[Row], field_list: list) -> None:
        writer = csv.DictWriter(fp, fieldnames=field_list)
        if field_list:
            writer.writeheader()
        for row in rows:
            writer.writerow(row)


class JSONLInputPlugin:
    name = "jsonl"
    extensions = [".jsonl"]
    typed = True

    @contextmanager
    def open(self, path: Path) -> Iterator[TableStream]:
        with path.open("r", encoding="utf-8") as fp:
            yield self._read(fp)

    @contextmanager
    def open_stream(self, stream: IO[str]) -> Iterator[TableStream]:
        yield self._read(stream)

    @staticmethod
    def _read(fp: IO[str]) -> TableStream:
        def generator():
            for line in fp:
                line = line.strip()
                if not line:
                    continue
                yield json.loads(line)

        iterator = generator()
        try:
            first = next(iterator)
        except StopIteration:
            return TableStream(rows=iter(()), fieldnames=[])

        def chain():
            yield first
            yield from iterator

        fieldnames = list(first.keys()) if isinstance(first, dict) else []
        return TableStream(rows=chain(), fieldnames=fieldnames)


class _JSONLRowAppender:
    def __init__(self, fp: IO[str]):
        self._fp = fp

    def append(self, row: Row) -> None:
        self._fp.write(json.dumps(row, ensure_ascii=False))
        self._fp.write("\n")

    def flush(self) -> None:
        self._fp.flush()


class JSONLOutputPlugin:
    name = "jsonl"
    extensions = [".jsonl"]

    def write(self, path: Path, rows: Iterable[Row], fieldnames: Sequence[str]) -> None:
        with path.open("w", encoding="utf-8") as fp:
            self._write_fp(fp, rows)

    def write_stream(
        self, stream: IO[str], rows: Iterable[Row], fieldnames: Sequence[str]
    ) -> None:
        self._write_fp(stream, rows)

    @contextmanager
    def open_append(
        self, path: Path, fieldnames: Sequence[str]
    ) -> Iterator[RowAppender]:
        fp = path.open("a", encoding="utf-8")
        try:
            yield _JSONLRowAppender(fp)
        finally:
            fp.close()

    @staticmethod
    def _write_fp(fp: IO[str], rows: Iterable[Row]) -> None:
        for row in rows:
            fp.write(json.dumps(row, ensure_ascii=False))
            fp.write("\n")


class XLSXInputPlugin:
    name = "xlsx"
    extensions = [".xlsx"]
    typed = True

    @contextmanager
    def open(self, path: Path) -> Iterator[TableStream]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if header is None:
                yield TableStream(rows=iter(()), fieldnames=[])
                return
            fieldnames = ["" if value is None else str(value) for value in header]

            def generator():
                for values in rows_iter:
                    row = {
                        fieldnames[i]: values[i] if i < len(values) else None
                        for i in range(len(fieldnames))
                    }
                    yield {
                        key: ("" if value is None else value)
                        for key, value in row.items()
                    }

            yield TableStream(rows=generator(), fieldnames=fieldnames)
        finally:
            workbook.close()


class XLSXOutputPlugin:
    name = "xlsx"
    extensions = [".xlsx"]

    def write(self, path: Path, rows: Iterable[Row], fieldnames: Sequence[str]) -> None:
        field_list = list(fieldnames)
        workbook = Workbook()
        sheet = workbook.active
        if field_list:
            sheet.append(list(field_list))
        for row in rows:
            sheet.append([row.get(field) for field in field_list])
        workbook.save(path)


def register_builtin_io(inputs: InputRegistry, outputs: OutputRegistry) -> None:
    inputs.register(CSVInputPlugin())
    outputs.register(CSVOutputPlugin())

    inputs.register(JSONLInputPlugin())
    outputs.register(JSONLOutputPlugin())

    inputs.register(XLSXInputPlugin())
    outputs.register(XLSXOutputPlugin())
