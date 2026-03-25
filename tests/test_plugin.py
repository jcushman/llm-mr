import csv
import json
from pathlib import Path

import llm
import pytest
from click.testing import CliRunner
from llm.cli import cli
from llm.plugins import pm
from openpyxl import Workbook, load_workbook

from llm_mr.plugin import register_commands


class MockModel(llm.Model):
    model_id = "demo"

    def __init__(self):
        self.responses = []
        self.last_prompt = None
        self.prompt_history = []
        self.last_schema = None
        self.schema_history = []

    def queue_response(self, text: str) -> None:
        self.responses.append(text)

    @property
    def supports_schema(self):
        return True

    def execute(self, prompt, stream, response, conversation):
        self.last_prompt = prompt
        self.prompt_history.append(prompt.prompt)
        if hasattr(prompt, "options") and "schema" in prompt.options:
            self.last_schema = prompt.options["schema"]
            self.schema_history.append(self.last_schema)
        text = self.responses.pop(0) if self.responses else ""
        return [text]


@pytest.fixture(scope="module", autouse=True)
def ensure_commands_registered():
    if "mr" not in cli.commands:
        register_commands(cli)
    yield


@pytest.fixture
def mock_model():
    model = MockModel()

    class TestPlugin:
        __name__ = "TestModelPlugin"

        @llm.hookimpl
        def register_models(self, register):
            register(model)

    plugin = TestPlugin()
    pm.register(plugin, name="llm-mr-test-model")
    try:
        yield model
    finally:
        pm.unregister(plugin=plugin)


def _write_csv(path: Path, rows):
    fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", encoding="utf-8", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _read_csv(path: Path):
    with path.open("r", encoding="utf-8", newline="") as fp:
        reader = csv.DictReader(fp)
        return list(reader)


def _write_jsonl(path: Path, rows):
    with path.open("w", encoding="utf-8") as fp:
        for row in rows:
            fp.write(json.dumps(row, ensure_ascii=False) + "\n")


def _read_jsonl(path: Path):
    rows = []
    with path.open("r", encoding="utf-8") as fp:
        for line in fp:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def _write_xlsx(path: Path, rows):
    wb = Workbook()
    ws = wb.active
    if rows:
        ws.append(list(rows[0].keys()))
        for row in rows:
            ws.append(list(row.values()))
    wb.save(path)


def _read_xlsx(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(v) for v in next(rows_iter)]
    result = []
    for values in rows_iter:
        result.append({header[i]: values[i] for i in range(len(header))})
    wb.close()
    return result


# ---------------------------------------------------------------------------
# Map tests
# ---------------------------------------------------------------------------


def test_map_prompt_mode(tmp_path, mock_model):
    """Map with -p treats instruction as a literal LLM prompt."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    rows = [
        {"id": "1", "status": "active", "note": "alpha", "mr_result": "existing"},
        {"id": "2", "status": "active", "note": "beta", "mr_result": ""},
        {"id": "3", "status": "inactive", "note": "gamma", "mr_result": ""},
    ]
    _write_csv(input_path, rows)

    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Processed beta"}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Return uppercase note",
            "-p",
            "-i",
            str(input_path),
            "--where",
            "status=active",
            "--few-shot",
            "1",
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output
    assert f"Wrote 3 rows to {output_path}" in result.output

    written_rows = _read_csv(output_path)
    assert written_rows[0]["mr_result"] == "existing"
    assert written_rows[1]["mr_result"] == "Processed beta"
    assert written_rows[2]["mr_result"] == ""

    prompt = mock_model.prompt_history[-1]
    assert "You are assisting with spreadsheet transformations." in prompt
    assert "<row_0>" in prompt
    assert "For each row, provide a single value for column 'mr_result'" in prompt


def test_map_expression_mode(tmp_path):
    """Map with -e treats instruction as a Python expression."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    rows = [
        {"id": "1", "name": "alice"},
        {"id": "2", "name": "bob"},
    ]
    _write_csv(input_path, rows)

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            'row["name"].upper()',
            "-e",
            "-i",
            str(input_path),
            "-c",
            "name_upper",
            "--output",
            str(output_path),
        ],
    )
    assert result.exit_code == 0, result.output
    assert f"Wrote 2 rows to {output_path}" in result.output

    written_rows = _read_csv(output_path)
    assert written_rows[0]["name_upper"] == "ALICE"
    assert written_rows[1]["name_upper"] == "BOB"


def test_map_expression_mode_with_builtins(tmp_path):
    """Expression mode has access to safe builtins."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    rows = [
        {"id": "1", "value": "42"},
        {"id": "2", "value": "7"},
    ]
    _write_csv(input_path, rows)

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            'str(int(row["value"]) * 2)',
            "-e",
            "-i",
            str(input_path),
            "-c",
            "doubled",
            "--output",
            str(output_path),
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_csv(output_path)
    assert written_rows[0]["doubled"] == "84"
    assert written_rows[1]["doubled"] == "14"


def test_map_cannot_use_both_p_and_e(tmp_path):
    """Using both -p and -e is an error."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    _write_csv(input_path, [{"id": "1"}])

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "some instruction",
            "-p",
            "-e",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
        ],
    )
    assert result.exit_code != 0
    assert "Cannot use both -p and -e" in result.output


def test_map_with_multiple_flag(tmp_path, mock_model):
    """--multiple flag emits multiple output rows per input row."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    rows = [
        {"id": "1", "text": "apple and banana"},
        {"id": "2", "text": "cherry"},
    ]
    _write_csv(input_path, rows)

    mock_model.queue_response(json.dumps({"row_0": {"fruit": ["apple", "banana"]}}))
    mock_model.queue_response(json.dumps({"row_0": {"fruit": ["cherry"]}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Extract individual fruit names from the text",
            "-p",
            "-i",
            str(input_path),
            "--column",
            "fruit",
            "--output",
            str(output_path),
            "--model",
            "demo",
            "--multiple",
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_csv(output_path)
    assert len(written_rows) == 3
    assert written_rows[0] == {"id": "1", "text": "apple and banana", "fruit": "apple"}
    assert written_rows[1] == {"id": "1", "text": "apple and banana", "fruit": "banana"}
    assert written_rows[2] == {"id": "2", "text": "cherry", "fruit": "cherry"}

    assert len(mock_model.prompt_history) == 2
    for prompt in mock_model.prompt_history:
        assert "For each row, provide zero or more values for column 'fruit'" in prompt


def test_map_with_multiple_flag_empty_list(tmp_path, mock_model):
    """--multiple with empty lists (zero results)."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    rows = [
        {"id": "1", "text": "nothing here"},
        {"id": "2", "text": "apple"},
    ]
    _write_csv(input_path, rows)

    mock_model.queue_response(json.dumps({"row_0": {"fruit": []}}))
    mock_model.queue_response(json.dumps({"row_0": {"fruit": ["apple"]}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Extract fruit names",
            "-p",
            "-i",
            str(input_path),
            "--column",
            "fruit",
            "--output",
            str(output_path),
            "--model",
            "demo",
            "--multiple",
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_csv(output_path)
    assert len(written_rows) == 1
    assert written_rows[0] == {"id": "2", "text": "apple", "fruit": "apple"}


def test_map_rejects_model_without_schema_support(tmp_path):
    """Map with -p fails with a clear error when model doesn't support schemas."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    _write_csv(input_path, [{"id": "1", "note": "test"}])

    class NoSchemaModel(llm.Model):
        model_id = "no-schema"

        def execute(self, prompt, stream, response, conversation):
            return ["result"]

    no_schema_model = NoSchemaModel()

    class TestPlugin:
        __name__ = "TestNoSchemaPlugin"

        @llm.hookimpl
        def register_models(self, register):
            register(no_schema_model)

    plugin = TestPlugin()
    pm.register(plugin, name="llm-mr-test-no-schema")

    try:
        runner = CliRunner()
        result = runner.invoke(
            cli,
            [
                "mr",
                "map",
                "Process note",
                "-p",
                "-i",
                str(input_path),
                "--output",
                str(output_path),
                "--model",
                "no-schema",
            ],
        )
        assert result.exit_code == 1
        assert "does not support schemas" in result.output
    finally:
        pm.unregister(plugin=plugin)


def test_map_writes_err_file_on_failure(tmp_path):
    """Failed map batches produce an .err sidecar with row indices."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    err_path = tmp_path / "output.csv.err"
    rows = [
        {"id": "1", "note": "alpha"},
        {"id": "2", "note": "beta"},
    ]
    _write_csv(input_path, rows)

    model = FailingMockModel(fail_on_calls={1})
    model.queue_response(json.dumps({"row_0": {"mr_result": "Result 1"}}))

    class TestPlugin:
        __name__ = "TestMapFailPlugin"

        @llm.hookimpl
        def register_models(self, register):
            register(model)

    plugin = TestPlugin()
    pm.register(plugin, name="llm-mr-test-map-fail")
    try:
        runner = CliRunner()
        result = runner.invoke(
            cli,
            [
                "mr",
                "map",
                "Process note",
                "-p",
                "-i",
                str(input_path),
                "--output",
                str(output_path),
                "--model",
                "failing-demo",
            ],
        )
        assert result.exit_code == 0
        assert "1 batches failed" in result.output

        written_rows = _read_csv(output_path)
        assert len(written_rows) == 2
        assert written_rows[0]["mr_result"] == "Result 1"
        assert written_rows[1]["mr_result"] == ""

        errors = _read_err(err_path)
        assert len(errors) == 1
        assert errors[0]["row_indices"] == [1]
        assert "Simulated API failure" in errors[0]["error"]

        assert "rerun to retry" in result.output
    finally:
        pm.unregister(plugin=plugin)


def test_map_resume_skips_done_rows(tmp_path, mock_model):
    """Re-running with existing output skips already-done rows."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"

    rows = [
        {"id": "1", "note": "alpha"},
        {"id": "2", "note": "beta"},
    ]
    _write_csv(input_path, rows)
    _write_csv(
        output_path,
        [
            {"id": "1", "note": "alpha", "mr_result": "Result 1"},
            {"id": "2", "note": "beta", "mr_result": ""},
        ],
    )

    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Result 2"}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_csv(output_path)
    assert len(written_rows) == 2
    assert written_rows[0]["mr_result"] == "Result 1"
    assert written_rows[1]["mr_result"] == "Result 2"

    assert len(mock_model.prompt_history) == 1


def test_map_worker_model(tmp_path, mock_model):
    """--worker-model is used for per-item LLM work."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    _write_csv(input_path, [{"id": "1", "note": "test"}])

    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "done"}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
            "--worker-model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output
    written = _read_csv(output_path)
    assert written[0]["mr_result"] == "done"


def test_map_expression_with_limit(tmp_path):
    """Map --limit processes only the first N rows."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    rows = [
        {"id": "1", "name": "alice"},
        {"id": "2", "name": "bob"},
        {"id": "3", "name": "carol"},
    ]
    _write_csv(input_path, rows)

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            'row["name"].upper()',
            "-e",
            "-i",
            str(input_path),
            "-c",
            "name_upper",
            "--limit",
            "2",
            "--output",
            str(output_path),
        ],
    )
    assert result.exit_code == 0, result.output
    assert "Wrote 2 rows" in result.output

    written_rows = _read_csv(output_path)
    assert len(written_rows) == 2
    assert written_rows[0]["name_upper"] == "ALICE"
    assert written_rows[1]["name_upper"] == "BOB"


def test_map_interactive_deterministic(tmp_path, mock_model):
    """Interactive mode (default) synthesizes and applies a deterministic expression."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    rows = [
        {"id": "1", "name": "alice"},
        {"id": "2", "name": "bob"},
    ]
    _write_csv(input_path, rows)

    mock_model.queue_response(
        json.dumps(
            {
                "expression": 'row["name"].upper()',
                "deterministic": True,
            }
        )
    )

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "uppercase the names",
            "-i",
            str(input_path),
            "-c",
            "name_upper",
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
        input="y\n",
    )
    assert result.exit_code == 0, result.output
    assert "Using deterministic expression" in result.output

    written_rows = _read_csv(output_path)
    assert written_rows[0]["name_upper"] == "ALICE"
    assert written_rows[1]["name_upper"] == "BOB"


def test_map_in_place(tmp_path):
    """--in-place writes results back to the input file."""
    input_path = tmp_path / "input.csv"
    rows = [
        {"id": "1", "name": "alice"},
        {"id": "2", "name": "bob"},
    ]
    _write_csv(input_path, rows)

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            'row["name"].upper()',
            "-e",
            "-i",
            str(input_path),
            "-c",
            "name_upper",
            "--in-place",
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_csv(input_path)
    assert len(written_rows) == 2
    assert written_rows[0]["name"] == "alice"
    assert written_rows[0]["name_upper"] == "ALICE"


# ---------------------------------------------------------------------------
# Reduce tests
# ---------------------------------------------------------------------------


def test_reduce_prompt_mode(tmp_path, mock_model):
    """Reduce with -p treats instruction as a literal LLM prompt."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    rows = [
        {"team": "A", "score": "10"},
        {"team": "A", "score": "12"},
        {"team": "B", "score": "5"},
    ]
    _write_csv(input_path, rows)

    mock_model.queue_response(json.dumps({"mr_result": "Summary A"}))
    mock_model.queue_response(json.dumps({"mr_result": "Summary B"}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "reduce",
            "Summarize group",
            "-p",
            "-i",
            str(input_path),
            "--group-by",
            "team",
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output
    assert f"Wrote 2 group results to {output_path}" in result.output

    written_rows = _read_csv(output_path)
    assert written_rows == [
        {"group": "A", "mr_result": "Summary A"},
        {"group": "B", "mr_result": "Summary B"},
    ]

    assert len(mock_model.prompt_history) == 2
    for prompt in mock_model.prompt_history:
        assert "You are summarizing a group of spreadsheet rows." in prompt


def test_reduce_group_key_column_option(tmp_path, mock_model):
    """--group-key-column renames the group key column in reduce output."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    rows = [
        {"team": "A", "score": "10"},
        {"team": "A", "score": "12"},
        {"team": "B", "score": "5"},
    ]
    _write_csv(input_path, rows)

    mock_model.queue_response(json.dumps({"summary": "Summary A"}))
    mock_model.queue_response(json.dumps({"summary": "Summary B"}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "reduce",
            "Summarize group",
            "-p",
            "-i",
            str(input_path),
            "--group-by",
            "team",
            "--group-key-column",
            "team",
            "-c",
            "summary",
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_csv(output_path)
    assert written_rows == [
        {"team": "A", "summary": "Summary A"},
        {"team": "B", "summary": "Summary B"},
    ]


def test_reduce_rejects_duplicate_key_and_result_column_names(tmp_path):
    """Reduce fails when --group-key-column matches -c."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    _write_csv(input_path, [{"team": "A", "score": "10"}])

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "reduce",
            "len(rows)",
            "-e",
            "-i",
            str(input_path),
            "--group-by",
            "team",
            "--group-key-column",
            "x",
            "-c",
            "x",
            "--output",
            str(output_path),
        ],
    )
    assert result.exit_code == 1
    assert "--group-key-column and --column must differ" in result.output


def test_reduce_expression_mode(tmp_path):
    """Reduce with -e uses a deterministic Python expression over rows."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    rows = [
        {"team": "A", "score": "10"},
        {"team": "A", "score": "12"},
        {"team": "B", "score": "5"},
    ]
    _write_csv(input_path, rows)

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "reduce",
            'sum(int(r["score"]) for r in rows)',
            "-e",
            "-i",
            str(input_path),
            "--group-by",
            "team",
            "--output",
            str(output_path),
        ],
    )
    assert result.exit_code == 0, result.output
    assert "Reduced 2 groups deterministically" in result.output

    written_rows = _read_csv(output_path)
    by_group = {r["group"]: r["mr_result"] for r in written_rows}
    assert by_group["A"] == "22"
    assert by_group["B"] == "5"


def test_reduce_validates_group_by_columns(tmp_path, mock_model):
    """Reduce fails when group-by column doesn't exist."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    _write_csv(input_path, [{"team": "A", "score": "10"}])

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "reduce",
            "Summarize",
            "-p",
            "-i",
            str(input_path),
            "--group-by",
            "does-not-exist",
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 1
    assert (
        "Column 'does-not-exist' specified in --group-by does not exist"
        in result.output
    )
    assert "Available columns: team, score" in result.output


def test_reduce_rejects_model_without_schema_support(tmp_path):
    """Reduce with -p fails when model doesn't support schemas."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    _write_csv(input_path, [{"team": "A", "score": "10"}])

    class NoSchemaModel(llm.Model):
        model_id = "no-schema-reduce"

        def execute(self, prompt, stream, response, conversation):
            return ["result"]

    no_schema_model = NoSchemaModel()

    class TestPlugin:
        __name__ = "TestNoSchemaPluginReduce"

        @llm.hookimpl
        def register_models(self, register):
            register(no_schema_model)

    plugin = TestPlugin()
    pm.register(plugin, name="llm-mr-test-no-schema-reduce")

    try:
        runner = CliRunner()
        result = runner.invoke(
            cli,
            [
                "mr",
                "reduce",
                "Summarize",
                "-p",
                "-i",
                str(input_path),
                "--group-by",
                "team",
                "--output",
                str(output_path),
                "--model",
                "no-schema-reduce",
            ],
        )
        assert result.exit_code == 1
        assert "does not support schemas" in result.output
    finally:
        pm.unregister(plugin=plugin)


def test_reduce_writes_err_file_on_failure(tmp_path):
    """Failed reduce groups produce an .err sidecar."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    err_path = tmp_path / "output.csv.err"
    rows = [
        {"team": "A", "score": "10"},
        {"team": "A", "score": "12"},
        {"team": "B", "score": "5"},
    ]
    _write_csv(input_path, rows)

    model = FailingMockModel(fail_on_calls={1})
    model.queue_response(json.dumps({"mr_result": "Summary A"}))

    class TestPlugin:
        __name__ = "TestFailPlugin"

        @llm.hookimpl
        def register_models(self, register):
            register(model)

    plugin = TestPlugin()
    pm.register(plugin, name="llm-mr-test-fail")
    try:
        runner = CliRunner()
        result = runner.invoke(
            cli,
            [
                "mr",
                "reduce",
                "Summarize group",
                "-p",
                "-i",
                str(input_path),
                "--group-by",
                "team",
                "--output",
                str(output_path),
                "--model",
                "failing-demo",
            ],
        )
        assert result.exit_code == 0
        assert "1/2 groups failed" in result.output

        written_rows = _read_csv(output_path)
        assert len(written_rows) == 1
        assert written_rows[0]["group"] == "A"
        assert written_rows[0]["mr_result"] == "Summary A"

        errors = _read_err(err_path)
        assert len(errors) == 1
        assert errors[0]["group_key"] == "B"
        assert "Simulated API failure" in errors[0]["error"]

        assert "rerun to retry" in result.output
    finally:
        pm.unregister(plugin=plugin)


def test_reduce_resume_skips_done_groups(tmp_path, mock_model):
    """Re-running with existing output skips already-done groups."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    rows = [
        {"team": "A", "score": "10"},
        {"team": "A", "score": "12"},
        {"team": "B", "score": "5"},
    ]
    _write_csv(input_path, rows)

    _write_csv(
        output_path,
        [
            {"group": "A", "mr_result": "Summary A"},
        ],
    )

    mock_model.queue_response(json.dumps({"mr_result": "Summary B"}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "reduce",
            "Summarize group",
            "-p",
            "-i",
            str(input_path),
            "--group-by",
            "team",
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_csv(output_path)
    by_group = {r["group"]: r["mr_result"] for r in written_rows}
    assert by_group["A"] == "Summary A"
    assert by_group["B"] == "Summary B"

    assert len(mock_model.prompt_history) == 1


def test_reduce_resume_all_done(tmp_path, mock_model):
    """Re-running when all groups are done is a no-op."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    _write_csv(input_path, [{"team": "A", "score": "10"}])
    _write_csv(
        output_path,
        [{"group": "A", "mr_result": "Summary A"}],
    )

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "reduce",
            "Summarize",
            "-p",
            "-i",
            str(input_path),
            "--group-by",
            "team",
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0
    assert "No groups required LLM processing" in result.output
    assert len(mock_model.prompt_history) == 0


# ---------------------------------------------------------------------------
# Filter tests
# ---------------------------------------------------------------------------


def test_filter_expression_mode(tmp_path):
    """Filter with -e uses a deterministic Python predicate."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    rows = [
        {"id": "1", "score": "15", "name": "alice"},
        {"id": "2", "score": "5", "name": "bob"},
        {"id": "3", "score": "20", "name": "carol"},
    ]
    _write_csv(input_path, rows)

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "filter",
            'int(row["score"]) >= 10',
            "-e",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
        ],
    )
    assert result.exit_code == 0, result.output
    assert "Kept 2/3 rows" in result.output

    written_rows = _read_csv(output_path)
    assert len(written_rows) == 2
    assert written_rows[0]["name"] == "alice"
    assert written_rows[1]["name"] == "carol"


def test_filter_expression_string_match(tmp_path):
    """Filter with -e can do string matching."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    rows = [
        {"id": "1", "topic": "prediction markets in 2024"},
        {"id": "2", "topic": "weather forecast"},
        {"id": "3", "topic": "polymarket prediction market"},
    ]
    _write_csv(input_path, rows)

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "filter",
            '"prediction market" in row["topic"].lower()',
            "-e",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
        ],
    )
    assert result.exit_code == 0, result.output
    assert "Kept 2/3 rows" in result.output

    written_rows = _read_csv(output_path)
    assert len(written_rows) == 2
    assert written_rows[0]["id"] == "1"
    assert written_rows[1]["id"] == "3"


def test_filter_prompt_mode(tmp_path, mock_model):
    """Filter with -p uses LLM to classify each row."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    rows = [
        {"id": "1", "topic": "machine learning"},
        {"id": "2", "topic": "gardening"},
        {"id": "3", "topic": "neural networks"},
    ]
    _write_csv(input_path, rows)

    mock_model.queue_response(json.dumps({"row_0": {"verdict": "keep"}}))
    mock_model.queue_response(json.dumps({"row_0": {"verdict": "discard"}}))
    mock_model.queue_response(json.dumps({"row_0": {"verdict": "keep"}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "filter",
            "about artificial intelligence",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output
    assert "Kept 2/3 rows" in result.output

    written_rows = _read_csv(output_path)
    assert len(written_rows) == 2
    assert written_rows[0]["topic"] == "machine learning"
    assert written_rows[1]["topic"] == "neural networks"

    for prompt in mock_model.prompt_history:
        assert "filter_criterion" in prompt


def test_filter_with_where_prefilter(tmp_path):
    """Filter --where applies before the instruction filter."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    rows = [
        {"id": "1", "status": "active", "score": "15"},
        {"id": "2", "status": "inactive", "score": "20"},
        {"id": "3", "status": "active", "score": "5"},
    ]
    _write_csv(input_path, rows)

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "filter",
            'int(row["score"]) >= 10',
            "-e",
            "-i",
            str(input_path),
            "--where",
            "status=active",
            "--output",
            str(output_path),
        ],
    )
    assert result.exit_code == 0, result.output
    assert "Kept 1/2 rows" in result.output

    written_rows = _read_csv(output_path)
    assert len(written_rows) == 1
    assert written_rows[0]["id"] == "1"


def test_filter_with_limit(tmp_path):
    """Filter --limit restricts rows before filtering."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    rows = [
        {"id": "1", "score": "15"},
        {"id": "2", "score": "5"},
        {"id": "3", "score": "20"},
    ]
    _write_csv(input_path, rows)

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "filter",
            'int(row["score"]) >= 10',
            "-e",
            "-i",
            str(input_path),
            "--limit",
            "2",
            "--output",
            str(output_path),
        ],
    )
    assert result.exit_code == 0, result.output
    assert "Kept 1/2 rows" in result.output

    written_rows = _read_csv(output_path)
    assert len(written_rows) == 1
    assert written_rows[0]["id"] == "1"


# ---------------------------------------------------------------------------
# Dry-run & verbose tests
# ---------------------------------------------------------------------------


def test_map_dry_run(tmp_path, mock_model):
    """--dry-run shows the sample prompt and schema without making LLM calls."""
    input_path = tmp_path / "input.csv"
    _write_csv(input_path, [{"id": "1", "note": "hello"}])

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Summarize the note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(tmp_path / "out.csv"),
            "--model",
            "demo",
            "--dry-run",
        ],
    )
    assert result.exit_code == 0, result.output
    assert "DRY RUN" in result.output
    assert "sample prompt" in result.output
    assert "spreadsheet_rows" in result.output
    assert "schema" in result.output
    assert "Would process" in result.output
    assert len(mock_model.prompt_history) == 0


def test_map_verbose(tmp_path, mock_model):
    """--verbose prints each prompt as it is sent."""
    input_path = tmp_path / "input.csv"
    _write_csv(input_path, [{"id": "1", "note": "hello"}])

    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "done"}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Summarize the note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(tmp_path / "out.csv"),
            "--model",
            "demo",
            "--verbose",
        ],
    )
    assert result.exit_code == 0, result.output
    assert "--- prompt ---" in result.output
    assert "spreadsheet_rows" in result.output
    assert "--- end prompt ---" in result.output
    assert len(mock_model.prompt_history) == 1


def test_map_log_tip(tmp_path, mock_model):
    """Map prints an llm logs tip after LLM processing."""
    input_path = tmp_path / "input.csv"
    _write_csv(input_path, [{"id": "1", "note": "hello"}])

    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "done"}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Summarize the note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(tmp_path / "out.csv"),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output
    assert "llm logs" in result.output


def test_reduce_dry_run(tmp_path, mock_model):
    """--dry-run on reduce shows sample prompt without making LLM calls."""
    input_path = tmp_path / "input.csv"
    _write_csv(
        input_path,
        [
            {"team": "A", "score": "10"},
            {"team": "A", "score": "12"},
        ],
    )

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "reduce",
            "Summarize group",
            "-p",
            "-i",
            str(input_path),
            "--group-by",
            "team",
            "--output",
            str(tmp_path / "out.csv"),
            "--model",
            "demo",
            "--dry-run",
        ],
    )
    assert result.exit_code == 0, result.output
    assert "DRY RUN" in result.output
    assert "Would process 1 groups" in result.output
    assert len(mock_model.prompt_history) == 0


def test_filter_dry_run(tmp_path, mock_model):
    """--dry-run on filter shows sample prompt without making LLM calls."""
    input_path = tmp_path / "input.csv"
    _write_csv(
        input_path,
        [
            {"id": "1", "topic": "machine learning"},
            {"id": "2", "topic": "gardening"},
        ],
    )

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "filter",
            "about AI",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(tmp_path / "out.csv"),
            "--model",
            "demo",
            "--dry-run",
        ],
    )
    assert result.exit_code == 0, result.output
    assert "DRY RUN" in result.output
    assert "filter_criterion" in result.output
    assert len(mock_model.prompt_history) == 0


# ---------------------------------------------------------------------------
# Format tests (JSONL, XLSX)
# ---------------------------------------------------------------------------


def test_map_expression_jsonl(tmp_path):
    """Map -e works with JSONL input and output."""
    input_path = tmp_path / "input.jsonl"
    output_path = tmp_path / "output.jsonl"
    _write_jsonl(
        input_path,
        [
            {"id": "1", "name": "alice"},
            {"id": "2", "name": "bob"},
        ],
    )

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            'row["name"].upper()',
            "-e",
            "-i",
            str(input_path),
            "-c",
            "name_upper",
            "--output",
            str(output_path),
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_jsonl(output_path)
    assert len(written_rows) == 2
    assert written_rows[0]["name_upper"] == "ALICE"
    assert written_rows[1]["name_upper"] == "BOB"


def test_map_expression_xlsx(tmp_path):
    """Map -e works with XLSX input and output."""
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    _write_xlsx(
        input_path,
        [
            {"id": "1", "name": "alice"},
            {"id": "2", "name": "bob"},
        ],
    )

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            'row["name"].upper()',
            "-e",
            "-i",
            str(input_path),
            "-c",
            "name_upper",
            "--output",
            str(output_path),
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_xlsx(output_path)
    assert len(written_rows) == 2
    assert written_rows[0]["name_upper"] == "ALICE"
    assert written_rows[1]["name_upper"] == "BOB"


# ---------------------------------------------------------------------------
# Piping tests
# ---------------------------------------------------------------------------


def test_filter_stdin_stdout_jsonl(tmp_path):
    """Filter reads JSONL from stdin and writes to stdout when no -i/-o."""
    rows = [
        {"id": "1", "score": "15"},
        {"id": "2", "score": "5"},
        {"id": "3", "score": "20"},
    ]
    stdin_data = "\n".join(json.dumps(r) for r in rows) + "\n"

    runner = CliRunner()
    result = runner.invoke(
        cli,
        ["mr", "filter", 'int(row["score"]) >= 10', "-e"],
        input=stdin_data,
    )
    assert result.exit_code == 0, result.stderr
    assert "Kept 2/3 rows" in result.stderr

    output_rows = [json.loads(line) for line in result.stdout.strip().splitlines()]
    assert len(output_rows) == 2
    assert output_rows[0]["id"] == "1"
    assert output_rows[1]["id"] == "3"


def test_map_stdin_stdout_jsonl(tmp_path):
    """Map reads JSONL from stdin and writes to stdout when no -i/-o."""
    rows = [
        {"id": "1", "name": "alice"},
        {"id": "2", "name": "bob"},
    ]
    stdin_data = "\n".join(json.dumps(r) for r in rows) + "\n"

    runner = CliRunner()
    result = runner.invoke(
        cli,
        ["mr", "map", 'row["name"].upper()', "-e", "-c", "upper"],
        input=stdin_data,
    )
    assert result.exit_code == 0, result.stderr

    output_rows = [json.loads(line) for line in result.stdout.strip().splitlines()]
    assert len(output_rows) == 2
    assert output_rows[0]["upper"] == "ALICE"
    assert output_rows[1]["upper"] == "BOB"


def test_reduce_stdin_stdout_jsonl(tmp_path):
    """Reduce reads JSONL from stdin and writes to stdout when no -i/-o."""
    rows = [
        {"team": "A", "score": "10"},
        {"team": "A", "score": "12"},
        {"team": "B", "score": "5"},
    ]
    stdin_data = "\n".join(json.dumps(r) for r in rows) + "\n"

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "reduce",
            'sum(int(r["score"]) for r in rows)',
            "-e",
            "--group-by",
            "team",
        ],
        input=stdin_data,
    )
    assert result.exit_code == 0, result.stderr
    assert "Reduced 2 groups deterministically" in result.stderr

    output_rows = [json.loads(line) for line in result.stdout.strip().splitlines()]
    by_group = {r["group"]: r["mr_result"] for r in output_rows}
    assert by_group["A"] == 22
    assert by_group["B"] == 5


def test_pipe_csv_with_format_flag(tmp_path):
    """Piping CSV via stdin/stdout using -f csv."""
    rows = [
        {"id": "1", "name": "alice"},
        {"id": "2", "name": "bob"},
    ]
    import io
    import csv as _csv

    buf = io.StringIO()
    writer = _csv.DictWriter(buf, fieldnames=["id", "name"])
    writer.writeheader()
    for r in rows:
        writer.writerow(r)
    stdin_data = buf.getvalue()

    runner = CliRunner()
    result = runner.invoke(
        cli,
        ["mr", "filter", 'row["name"] == "alice"', "-e", "-f", "csv"],
        input=stdin_data,
    )
    assert result.exit_code == 0, result.stderr

    reader = _csv.DictReader(io.StringIO(result.stdout))
    output_rows = list(reader)
    assert len(output_rows) == 1
    assert output_rows[0]["name"] == "alice"


def test_format_cascade_file_extension(tmp_path):
    """Format cascade: -i CSV input → output matches CSV when piping to stdout."""
    input_path = tmp_path / "input.csv"
    _write_csv(input_path, [{"id": "1", "name": "alice"}, {"id": "2", "name": "bob"}])

    runner = CliRunner()
    result = runner.invoke(
        cli,
        ["mr", "filter", 'row["name"] == "alice"', "-e", "-i", str(input_path)],
    )
    assert result.exit_code == 0, result.stderr

    import io
    import csv as _csv

    reader = _csv.DictReader(io.StringIO(result.stdout))
    output_rows = list(reader)
    assert len(output_rows) == 1
    assert output_rows[0]["name"] == "alice"


def test_format_cascade_output_format_override(tmp_path):
    """--output-format overrides the cascade."""
    input_path = tmp_path / "input.csv"
    _write_csv(input_path, [{"id": "1", "name": "alice"}])

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "filter",
            "True",
            "-e",
            "-i",
            str(input_path),
            "--output-format",
            "jsonl",
        ],
    )
    assert result.exit_code == 0, result.stderr

    output_rows = [json.loads(line) for line in result.stdout.strip().splitlines()]
    assert len(output_rows) == 1
    assert output_rows[0]["name"] == "alice"


def test_interactive_mode_blocked_with_stdin():
    """Interactive mode errors when reading from stdin (no -i)."""
    stdin_data = '{"id": "1"}\n'
    runner = CliRunner()
    result = runner.invoke(
        cli,
        ["mr", "filter", "some instruction"],
        input=stdin_data,
    )
    assert result.exit_code != 0
    assert "Cannot use interactive mode when reading from stdin" in result.output


def test_map_force_overwrites_existing(tmp_path, mock_model):
    """--force overwrites existing output that doesn't match input."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"

    _write_csv(input_path, [{"id": "1", "note": "alpha"}])
    _write_csv(output_path, [{"unrelated": "data"}])

    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Result 1"}}))

    runner = CliRunner()
    # Without --force, should error
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code != 0
    assert "does not match the input shape" in result.output

    # With --force, should succeed
    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Result 1"}}))
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
            "--model",
            "demo",
            "--force",
        ],
    )
    assert result.exit_code == 0, result.output
    written_rows = _read_csv(output_path)
    assert written_rows[0]["mr_result"] == "Result 1"


def test_map_err_flag_overrides_default_sidecar(tmp_path):
    """--err overrides the default <output>.err location."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    custom_err = tmp_path / "custom_errors.jsonl"
    default_err = tmp_path / "output.csv.err"
    rows = [
        {"id": "1", "note": "alpha"},
        {"id": "2", "note": "beta"},
    ]
    _write_csv(input_path, rows)

    model = FailingMockModel(fail_on_calls={1})
    model.queue_response(json.dumps({"row_0": {"mr_result": "Result 1"}}))

    class TestPlugin:
        __name__ = "TestErrFlagPlugin"

        @llm.hookimpl
        def register_models(self, register):
            register(model)

    plugin = TestPlugin()
    pm.register(plugin, name="llm-mr-test-err-flag")
    try:
        runner = CliRunner()
        result = runner.invoke(
            cli,
            [
                "mr",
                "map",
                "Process note",
                "-p",
                "-i",
                str(input_path),
                "--output",
                str(output_path),
                "--model",
                "failing-demo",
                "--err",
                str(custom_err),
            ],
        )
        assert result.exit_code == 0
        assert "1 batches failed" in result.output

        assert not default_err.exists(), "Default .err should not be created"
        errors = _read_err(custom_err)
        assert len(errors) == 1
        assert errors[0]["row_indices"] == [1]
    finally:
        pm.unregister(plugin=plugin)


def test_map_err_flag_enables_file_logging_with_stdout(tmp_path):
    """--err enables file-based error logging when output goes to stdout."""
    rows = [
        {"id": "1", "note": "alpha"},
        {"id": "2", "note": "beta"},
    ]
    stdin_data = "\n".join(json.dumps(r) for r in rows) + "\n"
    custom_err = tmp_path / "errors.jsonl"

    model = FailingMockModel(fail_on_calls={1})
    model.queue_response(json.dumps({"row_0": {"mr_result": "Result 1"}}))

    class TestPlugin:
        __name__ = "TestErrStdoutPlugin"

        @llm.hookimpl
        def register_models(self, register):
            register(model)

    plugin = TestPlugin()
    pm.register(plugin, name="llm-mr-test-err-stdout")
    try:
        runner = CliRunner()
        result = runner.invoke(
            cli,
            [
                "mr",
                "map",
                "Process note",
                "-p",
                "--model",
                "failing-demo",
                "--err",
                str(custom_err),
            ],
            input=stdin_data,
        )
        assert result.exit_code == 0

        errors = _read_err(custom_err)
        assert len(errors) == 1
        assert errors[0]["row_indices"] == [1]
        assert "row_indices" not in result.stderr
    finally:
        pm.unregister(plugin=plugin)


def test_map_resume_with_limit_expansion(tmp_path, mock_model):
    """Running with -n 5 then without -n processes remaining rows."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"

    rows = [
        {"id": "1", "note": "alpha"},
        {"id": "2", "note": "beta"},
        {"id": "3", "note": "gamma"},
    ]
    _write_csv(input_path, rows)

    # First run: -n 1 processes only first row
    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Result 1"}}))
    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
            "--model",
            "demo",
            "--limit",
            "1",
        ],
    )
    assert result.exit_code == 0, result.output
    assert "Wrote 1 rows" in result.output

    # Second run: processes remaining rows
    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Result 2"}}))
    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Result 3"}}))
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_csv(output_path)
    assert len(written_rows) == 3
    assert written_rows[0]["mr_result"] == "Result 1"
    assert written_rows[1]["mr_result"] == "Result 2"
    assert written_rows[2]["mr_result"] == "Result 3"


def test_err_records_to_stderr_when_stdout_output(tmp_path):
    """When output is stdout, error records go to stderr instead of sidecar."""
    rows = [
        {"id": "1", "note": "alpha"},
        {"id": "2", "note": "beta"},
    ]
    stdin_data = "\n".join(json.dumps(r) for r in rows) + "\n"

    model = FailingMockModel(fail_on_calls={1})
    model.queue_response(json.dumps({"row_0": {"mr_result": "Result 1"}}))

    class TestPlugin:
        __name__ = "TestStderrErrPlugin"

        @llm.hookimpl
        def register_models(self, register):
            register(model)

    plugin = TestPlugin()
    pm.register(plugin, name="llm-mr-test-stderr-err")
    try:
        runner = CliRunner()
        result = runner.invoke(
            cli,
            ["mr", "map", "Process note", "-p", "--model", "failing-demo"],
            input=stdin_data,
        )
        assert result.exit_code == 0
        assert "Simulated API failure" in result.stderr
        assert "row_indices" in result.stderr
    finally:
        pm.unregister(plugin=plugin)


def test_in_place_requires_input_file():
    """--in-place without -i is an error."""
    runner = CliRunner()
    result = runner.invoke(
        cli,
        ["mr", "map", "some instruction", "-e", "--in-place"],
        input='{"id": "1"}\n',
    )
    assert result.exit_code != 0
    assert "--in-place requires -i" in result.output


# ---------------------------------------------------------------------------
# Protocol and external plugin tests
# ---------------------------------------------------------------------------


def test_streamable_protocol_checks():
    """Builtin plugins satisfy the correct streaming protocols."""
    from llm_mr.io_plugins import (
        CSVInputPlugin,
        CSVOutputPlugin,
        JSONLInputPlugin,
        JSONLOutputPlugin,
        XLSXInputPlugin,
        XLSXOutputPlugin,
    )
    from llm_mr.registries import AppendableOutput, StreamableInput, StreamableOutput

    assert isinstance(CSVInputPlugin(), StreamableInput)
    assert isinstance(CSVOutputPlugin(), StreamableOutput)
    assert isinstance(CSVOutputPlugin(), AppendableOutput)
    assert isinstance(JSONLInputPlugin(), StreamableInput)
    assert isinstance(JSONLOutputPlugin(), StreamableOutput)
    assert isinstance(JSONLOutputPlugin(), AppendableOutput)
    assert not isinstance(XLSXInputPlugin(), StreamableInput)
    assert not isinstance(XLSXOutputPlugin(), StreamableOutput)
    assert not isinstance(XLSXOutputPlugin(), AppendableOutput)


def test_non_streamable_input_temp_file_fallback(monkeypatch):
    """Piping to a Path-only input plugin spools through a temp file."""
    import io
    from contextlib import contextmanager

    from llm_mr.processors import _open_input
    from llm_mr.registries import (
        InputRegistry,
        OutputRegistry,
        PluginContext,
        TableStream,
    )

    class PathOnlyInput:
        name = "pathonly"
        extensions = [".pathonly"]

        @contextmanager
        def open(self, path):
            with path.open("r") as fp:
                rows = [json.loads(line) for line in fp if line.strip()]
            fieldnames = list(rows[0].keys()) if rows else []
            yield TableStream(rows=iter(rows), fieldnames=fieldnames)

    inputs = InputRegistry()
    inputs.register(PathOnlyInput())
    context = PluginContext(inputs=inputs, outputs=OutputRegistry())

    fake_stdin = io.StringIO(
        '{"id": "1", "name": "alice"}\n{"id": "2", "name": "bob"}\n'
    )
    monkeypatch.setattr("sys.stdin", fake_stdin)

    with _open_input(context, None, "pathonly") as stream:
        rows = list(stream.rows)

    assert len(rows) == 2
    assert rows[0] == {"id": "1", "name": "alice"}
    assert rows[1] == {"id": "2", "name": "bob"}


def test_non_streamable_output_temp_file_fallback(monkeypatch):
    """Writing to stdout with a Path-only output plugin goes through a temp file."""
    import io

    from llm_mr.processors import _output_writer
    from llm_mr.registries import InputRegistry, OutputRegistry, PluginContext

    class PathOnlyOutput:
        name = "pathonly"
        extensions = [".pathonly"]

        def write(self, path, rows, fieldnames):
            with path.open("w", encoding="utf-8") as fp:
                for row in rows:
                    fp.write(json.dumps(row) + "\n")

    outputs = OutputRegistry()
    outputs.register(PathOnlyOutput())
    context = PluginContext(inputs=InputRegistry(), outputs=outputs)

    fake_buffer = io.BytesIO()

    class FakeStdout:
        buffer = fake_buffer

    monkeypatch.setattr("sys.stdout", FakeStdout())

    write = _output_writer(context, None, "pathonly")
    write([{"id": "1"}, {"id": "2"}], ["id"])

    output = fake_buffer.getvalue().decode()
    rows = [json.loads(line) for line in output.strip().splitlines()]
    assert len(rows) == 2
    assert rows[0] == {"id": "1"}
    assert rows[1] == {"id": "2"}


def test_hook_registration():
    """Third-party plugins register via the hook system."""
    from contextlib import contextmanager

    from llm_mr.hookspecs import mr_hookimpl, mr_pm
    from llm_mr.registries import InputRegistry, TableStream

    class FakePlugin:
        name = "hooktest"
        extensions = [".hooktest"]

        @contextmanager
        def open(self, path):
            yield TableStream(rows=iter(()), fieldnames=[])

    class HookImpl:
        __name__ = "TestHookImpl"

        @mr_hookimpl
        def register_mr_inputs(self, register):
            register(FakePlugin())

    impl = HookImpl()
    mr_pm.register(impl)
    try:
        registry = InputRegistry()
        mr_pm.hook.register_mr_inputs(register=registry.register)
        plugin = registry.for_name("hooktest")
        assert plugin.name == "hooktest"
        assert plugin.extensions == [".hooktest"]
    finally:
        mr_pm.unregister(impl)


# ---------------------------------------------------------------------------
# Resume and WAL tests
# ---------------------------------------------------------------------------


def test_map_resume_jsonl(tmp_path, mock_model):
    """Map resume works with JSONL format."""
    input_path = tmp_path / "input.jsonl"
    output_path = tmp_path / "output.jsonl"

    _write_jsonl(
        input_path,
        [
            {"id": "1", "note": "alpha"},
            {"id": "2", "note": "beta"},
            {"id": "3", "note": "gamma"},
        ],
    )
    _write_jsonl(
        output_path,
        [
            {"id": "1", "note": "alpha", "mr_result": "Result 1"},
        ],
    )

    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Result 2"}}))
    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Result 3"}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_jsonl(output_path)
    assert len(written_rows) == 3
    assert written_rows[0]["mr_result"] == "Result 1"
    assert written_rows[1]["mr_result"] == "Result 2"
    assert written_rows[2]["mr_result"] == "Result 3"

    # Only 2 LLM calls (rows 2 and 3), not 3
    assert len(mock_model.prompt_history) == 2


def test_map_resume_incremental_append(tmp_path, mock_model):
    """When existing output has only done rows, new rows are appended."""
    input_path = tmp_path / "input.jsonl"
    output_path = tmp_path / "output.jsonl"

    _write_jsonl(
        input_path,
        [
            {"id": "1", "note": "alpha"},
            {"id": "2", "note": "beta"},
        ],
    )
    _write_jsonl(
        output_path,
        [
            {"id": "1", "note": "alpha", "mr_result": "Result 1"},
        ],
    )

    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Result 2"}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_jsonl(output_path)
    assert len(written_rows) == 2
    assert written_rows[0]["mr_result"] == "Result 1"
    assert written_rows[1]["mr_result"] == "Result 2"

    # WAL should be cleaned up
    wal_path = Path(str(output_path) + ".wal")
    assert not wal_path.exists()


def test_map_incremental_new_file(tmp_path, mock_model):
    """Fresh run to a new JSONL file uses incremental append (ADR Case 1)."""
    input_path = tmp_path / "input.jsonl"
    output_path = tmp_path / "output.jsonl"

    _write_jsonl(
        input_path,
        [
            {"id": "1", "note": "alpha"},
            {"id": "2", "note": "beta"},
        ],
    )

    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Result 1"}}))
    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Result 2"}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_jsonl(output_path)
    assert len(written_rows) == 2
    assert written_rows[0]["mr_result"] == "Result 1"
    assert written_rows[1]["mr_result"] == "Result 2"

    # WAL should be cleaned up
    wal_path = Path(str(output_path) + ".wal")
    assert not wal_path.exists()


def test_map_resume_all_done(tmp_path, mock_model):
    """When all rows are already done, no LLM calls are made."""
    input_path = tmp_path / "input.jsonl"
    output_path = tmp_path / "output.jsonl"

    _write_jsonl(
        input_path,
        [
            {"id": "1", "note": "alpha"},
            {"id": "2", "note": "beta"},
        ],
    )
    _write_jsonl(
        output_path,
        [
            {"id": "1", "note": "alpha", "mr_result": "Result 1"},
            {"id": "2", "note": "beta", "mr_result": "Result 2"},
        ],
    )

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output
    assert "No rows required LLM processing" in result.output
    assert len(mock_model.prompt_history) == 0


def test_map_wal_survives_failure(tmp_path):
    """WAL records survive batch failures and enable resume."""
    input_path = tmp_path / "input.jsonl"
    output_path = tmp_path / "output.jsonl"

    _write_jsonl(
        input_path,
        [
            {"id": "1", "note": "alpha"},
            {"id": "2", "note": "beta"},
            {"id": "3", "note": "gamma"},
        ],
    )

    model = FailingMockModel(fail_on_calls={1})
    model.queue_response(json.dumps({"row_0": {"mr_result": "Result 1"}}))
    model.queue_response(json.dumps({"row_0": {"mr_result": "Result 3"}}))

    class TestPlugin:
        __name__ = "TestWalFailPlugin"

        @llm.hookimpl
        def register_models(self, register):
            register(model)

    plugin = TestPlugin()
    pm.register(plugin, name="llm-mr-test-wal-fail")
    try:
        runner = CliRunner()
        result = runner.invoke(
            cli,
            [
                "mr",
                "map",
                "Process note",
                "-p",
                "-i",
                str(input_path),
                "--output",
                str(output_path),
                "--model",
                "failing-demo",
            ],
        )
        assert result.exit_code == 0
        assert "1 batches failed" in result.output

        written_rows = _read_jsonl(output_path)
        assert len(written_rows) == 3
        assert written_rows[0]["mr_result"] == "Result 1"
        assert written_rows[1].get("mr_result", "") == ""
        assert written_rows[2]["mr_result"] == "Result 3"
    finally:
        pm.unregister(plugin=plugin)


def test_wal_tolerates_corrupt_lines(tmp_path):
    """_read_wal and _read_errors skip truncated/corrupt JSON lines."""
    from llm_mr.processors import _read_errors, _read_wal

    wal_path = tmp_path / "test.wal"
    wal_path.write_text(
        '{"i": 0, "c": "mr_result", "v": "ok"}\n'
        '{"i": 1, "c": "mr_re\n'  # truncated mid-write
        '{"i": 2, "c": "mr_result", "v": "also ok"}\n'
    )
    records = _read_wal(wal_path)
    assert len(records) == 2
    assert records[0]["i"] == 0
    assert records[1]["i"] == 2

    err_path = tmp_path / "test.err"
    err_path.write_text(
        '{"row_indices": [0], "error": "boom"}\n'
        "not json at all\n"
        '{"row_indices": [1], "error": "bang"}\n'
    )
    errors = _read_errors(err_path)
    assert len(errors) == 2
    assert errors[0]["error"] == "boom"
    assert errors[1]["error"] == "bang"


def test_map_force_flag(tmp_path, mock_model):
    """--force overwrites an existing output file that doesn't match."""
    input_path = tmp_path / "input.jsonl"
    output_path = tmp_path / "output.jsonl"

    _write_jsonl(input_path, [{"id": "1", "note": "alpha"}])
    _write_jsonl(output_path, [{"completely": "different", "data": "here"}])

    # Without --force: error
    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code != 0
    assert "does not match the input shape" in result.output

    # With --force: success
    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Result 1"}}))
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
            "--model",
            "demo",
            "--force",
        ],
    )
    assert result.exit_code == 0, result.output
    written_rows = _read_jsonl(output_path)
    assert written_rows[0]["mr_result"] == "Result 1"


def test_reduce_resume_with_partial_output(tmp_path, mock_model):
    """Reduce resume skips already-done groups."""
    input_path = tmp_path / "input.jsonl"
    output_path = tmp_path / "output.jsonl"

    _write_jsonl(
        input_path,
        [
            {"team": "A", "score": "10"},
            {"team": "A", "score": "12"},
            {"team": "B", "score": "5"},
            {"team": "C", "score": "8"},
        ],
    )
    _write_jsonl(
        output_path,
        [
            {"group": "A", "mr_result": "Summary A"},
        ],
    )

    mock_model.queue_response(json.dumps({"mr_result": "Summary B"}))
    mock_model.queue_response(json.dumps({"mr_result": "Summary C"}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "reduce",
            "Summarize group",
            "-p",
            "-i",
            str(input_path),
            "--group-by",
            "team",
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_jsonl(output_path)
    by_group = {r["group"]: r["mr_result"] for r in written_rows}
    assert by_group["A"] == "Summary A"
    assert by_group["B"] == "Summary B"
    assert by_group["C"] == "Summary C"

    assert len(mock_model.prompt_history) == 2


def test_map_resume_with_where(tmp_path, mock_model):
    """Resume works correctly with --where filter."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"

    rows = [
        {"id": "1", "status": "active", "note": "alpha"},
        {"id": "2", "status": "inactive", "note": "beta"},
        {"id": "3", "status": "active", "note": "gamma"},
    ]
    _write_csv(input_path, rows)
    _write_csv(
        output_path,
        [
            {"id": "1", "status": "active", "note": "alpha", "mr_result": "Result 1"},
            {"id": "2", "status": "inactive", "note": "beta", "mr_result": ""},
            {"id": "3", "status": "active", "note": "gamma", "mr_result": ""},
        ],
    )

    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Result 3"}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process note",
            "-p",
            "-i",
            str(input_path),
            "--where",
            "status=active",
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_csv(output_path)
    assert len(written_rows) == 3
    assert written_rows[0]["mr_result"] == "Result 1"
    assert written_rows[1]["mr_result"] == ""
    assert written_rows[2]["mr_result"] == "Result 3"

    assert len(mock_model.prompt_history) == 1


def test_map_no_resume_on_stdout(tmp_path, mock_model):
    """Stdout output never resumes — always processes all rows."""
    rows = [
        {"id": "1", "note": "alpha"},
        {"id": "2", "note": "beta"},
    ]
    stdin_data = "\n".join(json.dumps(r) for r in rows) + "\n"

    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "R1"}}))
    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "R2"}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        ["mr", "map", "Process note", "-p", "--model", "demo"],
        input=stdin_data,
    )
    assert result.exit_code == 0, result.stderr
    assert len(mock_model.prompt_history) == 2


def test_map_resume_cleans_up_wal(tmp_path, mock_model):
    """WAL is deleted after successful completion."""
    input_path = tmp_path / "input.jsonl"
    output_path = tmp_path / "output.jsonl"
    wal_path = Path(str(output_path) + ".wal")

    _write_jsonl(input_path, [{"id": "1", "note": "alpha"}])

    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Result 1"}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output
    assert not wal_path.exists()


def test_reduce_force_flag(tmp_path, mock_model):
    """--force on reduce overwrites existing output."""
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"

    _write_csv(input_path, [{"team": "A", "score": "10"}])
    _write_csv(output_path, [{"group": "A", "mr_result": "Old summary"}])

    mock_model.queue_response(json.dumps({"mr_result": "New summary"}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "reduce",
            "Summarize group",
            "-p",
            "-i",
            str(input_path),
            "--group-by",
            "team",
            "--output",
            str(output_path),
            "--model",
            "demo",
            "--force",
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_csv(output_path)
    assert written_rows[0]["mr_result"] == "New summary"


def test_map_resume_wal_recovery(tmp_path, mock_model):
    """WAL entries from a previous interrupted run are picked up on resume."""
    input_path = tmp_path / "input.jsonl"
    output_path = tmp_path / "output.jsonl"
    wal_path = Path(str(output_path) + ".wal")

    _write_jsonl(
        input_path,
        [
            {"id": "1", "note": "alpha"},
            {"id": "2", "note": "beta"},
            {"id": "3", "note": "gamma"},
        ],
    )
    # Output has first row done
    _write_jsonl(
        output_path,
        [
            {"id": "1", "note": "alpha", "mr_result": "Result 1"},
        ],
    )
    # WAL has second row done (from interrupted run)
    with wal_path.open("w", encoding="utf-8") as fp:
        fp.write(json.dumps({"i": 1, "c": "mr_result", "v": "Result 2"}) + "\n")

    # Only row 3 should need processing
    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Result 3"}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_jsonl(output_path)
    assert len(written_rows) == 3
    assert written_rows[0]["mr_result"] == "Result 1"
    assert written_rows[1]["mr_result"] == "Result 2"
    assert written_rows[2]["mr_result"] == "Result 3"

    assert len(mock_model.prompt_history) == 1
    assert not wal_path.exists()


def test_map_resume_in_place_skips_done(tmp_path, mock_model):
    """--in-place with existing target column values skips done rows."""
    input_path = tmp_path / "input.csv"
    rows = [
        {"id": "1", "name": "alice", "mr_result": "existing"},
        {"id": "2", "name": "bob", "mr_result": ""},
    ]
    _write_csv(input_path, rows)

    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "new"}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process name",
            "-p",
            "-i",
            str(input_path),
            "--in-place",
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output

    written_rows = _read_csv(input_path)
    assert written_rows[0]["mr_result"] == "existing"
    assert written_rows[1]["mr_result"] == "new"
    assert len(mock_model.prompt_history) == 1


def test_superset_matching():
    """Unit test for _is_superset helper."""
    from llm_mr.processors import _is_superset

    assert _is_superset(
        {"id": "1", "name": "alice", "mr_result": "done"},
        {"id": "1", "name": "alice"},
    )
    assert not _is_superset(
        {"id": "1", "name": "alice"},
        {"id": "1", "name": "alice", "extra": "field"},
    )
    assert not _is_superset(
        {"id": "1", "name": "bob", "mr_result": "done"},
        {"id": "1", "name": "alice"},
    )
    # str/int cross-type: matches (CSV round-trip tolerance)
    assert _is_superset({"count": "1"}, {"count": 1})
    assert _is_superset({"count": 1}, {"count": "1"})
    # Same-type mismatch: does not match
    assert not _is_superset({"count": 1}, {"count": 2})
    assert not _is_superset({"count": "1"}, {"count": "2"})
    # int/int same-type equality
    assert _is_superset({"count": 1, "extra": "x"}, {"count": 1})
    # None vs "None": both strings → no false match; different types → no coercion
    assert not _is_superset({"v": None}, {"v": "None"})


def test_map_resume_state():
    """Unit test for _match_map_output."""
    from llm_mr.processors import _match_map_output

    input_rows = [
        {"id": "1", "name": "alice"},
        {"id": "2", "name": "bob"},
        {"id": "3", "name": "carol"},
    ]
    output_rows = [
        {"id": "1", "name": "alice", "result": "A"},
    ]
    state = _match_map_output(input_rows, output_rows, "result")
    assert state.done_indices == {0}
    assert state.gap_indices == []
    assert state.tail_start == 1


def test_reorder_buffer():
    """Unit test for ReorderBuffer."""
    from llm_mr.processors import ReorderBuffer

    buf = ReorderBuffer()

    # Add batch 2 first (out of order)
    ready = buf.add_batch(2, [(20, {"id": "20"})])
    assert ready == []

    # Add batch 0
    ready = buf.add_batch(0, [(0, {"id": "0"})])
    assert len(ready) == 1
    assert ready[0][0] == 0

    # Add batch 1 — should flush both batch 1 and buffered batch 2
    ready = buf.add_batch(1, [(10, {"id": "10"})])
    assert len(ready) == 2
    assert ready[0][0] == 10
    assert ready[1][0] == 20


def test_map_resume_empty_output_file(tmp_path, mock_model):
    """Empty output file is treated as fresh start (no resume)."""
    input_path = tmp_path / "input.jsonl"
    output_path = tmp_path / "output.jsonl"

    _write_jsonl(input_path, [{"id": "1", "note": "alpha"}])
    output_path.touch()

    mock_model.queue_response(json.dumps({"row_0": {"mr_result": "Result 1"}}))

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "mr",
            "map",
            "Process note",
            "-p",
            "-i",
            str(input_path),
            "--output",
            str(output_path),
            "--model",
            "demo",
        ],
    )
    assert result.exit_code == 0, result.output
    written_rows = _read_jsonl(output_path)
    assert len(written_rows) == 1
    assert written_rows[0]["mr_result"] == "Result 1"


# ---------------------------------------------------------------------------
# Shared test helpers
# ---------------------------------------------------------------------------


class FailingMockModel(llm.Model):
    """Model that fails on specific calls by index."""

    model_id = "failing-demo"

    def __init__(self, fail_on_calls=None):
        self.fail_on_calls = fail_on_calls or set()
        self.responses = []
        self.call_count = 0
        self.last_prompt = None
        self.prompt_history = []

    def queue_response(self, text: str) -> None:
        self.responses.append(text)

    @property
    def supports_schema(self):
        return True

    def execute(self, prompt, stream, response, conversation):
        idx = self.call_count
        self.call_count += 1
        self.last_prompt = prompt
        self.prompt_history.append(prompt.prompt)
        if idx in self.fail_on_calls:
            raise RuntimeError(f"Simulated API failure on call {idx}")
        text = self.responses.pop(0) if self.responses else ""
        return [text]


def _read_err(path):
    if not path.exists():
        return []
    records = []
    with path.open("r", encoding="utf-8") as fp:
        for line in fp:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    return records
